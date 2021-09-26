import tkinter as tk
import tkinter.filedialog as fd
from tkinter import ttk
import os
import sys
#import PyPDF2
import fitz
import pytesseract
from PIL import Image
from openpyxl import load_workbook
import pandas as pd
import requests
import shutil
from pathlib import Path

def dir_cls():
    folder = 'c:/bu/tmp'
    for the_file in os.listdir(folder):
        file_path = os.path.join(folder, the_file)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
        except Exception as e:
            print(e)

def concate_exel():
    path = Path('c:/bu/tmp/')
    min_excel_file_size = 100
    df = pd.concat([pd.read_excel(f) 
                    for f in path.glob("*.xlsx") 
                    if f.stat().st_size >= min_excel_file_size],
                   ignore_index=True)

    df.to_excel('c:/bu/final.xlsx')    

class App(tk.Tk):

    def __init__(self):
        super().__init__()
        global  txt0,txt1
        lbl0 = tk.Label(self, text="Программа скачивает pdf файлы и распазнает их")  
        lbl0.pack(padx=6, pady=1)  
        lbl1 = tk.Label(self, text="Результат заносится в файл c:/bu/resultat.xlsx")  
        lbl1.pack(padx=6, pady=1) 
        lbl1 = tk.Label(self, text="Укажите колонку с номером БУ")  
        lbl1.pack(padx=6, pady=1)
        kolonki = ('A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T')
        txt0 = ttk.Combobox(self)
        txt0['values'] = kolonki
        #txt0.set('D')    
        txt0.pack(padx=6, pady=1)
        lbl1 = tk.Label(self, text="Укажите колонку с ссылкой на файл")  
        lbl1.pack(padx=6, pady=1) 
        txt1 = ttk.Combobox(self)
        txt1['values'] = kolonki
        #txt1.set('I')        
        txt1.pack(padx=6, pady=1) 
        txt0.bind("<<ComboboxSelected>>",choose_rb)
        txt1.bind("<<ComboboxSelected>>",choose_rb)
        
        btn_file = tk.Button(self, text="Выбрать файл",
                             command=self.choose_file)
        btn_file.pack(padx=60, pady=10)
        btn_exit =tk.Button(self, text ="      Выход       ", command=self.onExit)
        btn_exit.pack(padx=60, pady=10)

    def choose_file(self):
        filetypes = (("Exel файл", "*.xlsx"),
                     ("PDF файл", "*.pdf"),
                     ("Любой", "*"))
        filename = fd.askopenfilename(title="Открыть файл", initialdir="/",                                 # получили имя файла с путем
                                      filetypes=filetypes)
                                      
        if filename:
            print(filename)
            # работа с экселем
            
            tab_link = []                                                                                      # список строк для будущей таблицы   
            wb = load_workbook(filename)                                                                       # файл с ссылками на сканы актов возврата БУ
            ws = wb.active                                                                                     # выбор активной страницы
            headers = {
                        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/12.0.3 Safari/605.1.15",
                        "Accept-Language": "en"
                    }
            for i in range(4, ws.max_row + 1):                                                              # перебираем строки исх таблицы
                print("обрабатываем  "  +str(i-3)+ "  из "+ str(ws.max_row -3))
                row_links =[] # строка для будущей таблицы    
                nom_bu = (ws.cell(row=i, column=(int(rb)+1)).value)                                                   # получаем номер БУ 
                adressURL=(ws.cell(row=i, column=(int(ra)+1)).value)                # получаем ссылку
                if len(nom_bu)!=9:
                    print('ошибка выбора колонки!')
                    break
                if('http' in str(adressURL)):    # возможно  это даже ссылка
                    
                    send=requests.get(adressURL, headers=headers)                                               # загружаем файл в объект
                    name_of_PDF_file ='c:/bu/tmp/' +str(nom_bu)+'.pdf'
                    #print(name_of_PDF_file)
                    if send.status_code == 200:                                                                 # если успешно  записывваем объект как файл pdf
                        #print('Success!')
                        row_links.append(nom_bu)                                                                #в список добавляем номер БУ
                        row_links.append(adressURL)                                                             #в список добавляемссылку на скан акта
                        with open(name_of_PDF_file,"wb")as code:
                            code.write(send.content)                                                                #записываем результат  в папку tmp
                        # работа с pdf
                        pdf_file = fitz.open(name_of_PDF_file)                                                      #  открываем скачаный файл
                        location = 'c:/bu/tmp/'                                                                     #  папка для текстовых файлов
                        number_of_pages = len(pdf_file)
                        row_links.append(number_of_pages) #
                        #подсчет количества страниц в pdf
                        
                        file_name  =1
                        for current_page_index in range(1):                     #итерация по каждой странице в pdf number_of_pages
                            for img_index,img in enumerate(pdf_file.getPageImageList(current_page_index)):          # итерация по каждому изображению на каждой странице PDF
                                if img_index>10:
                                    break  #  игнорим 11 страницу
                                xref = img[0]
                                image = fitz.Pixmap(pdf_file, xref)
                                img_filename = name_of_PDF_file+'-'+str(current_page_index)+'-'+str(img_index) + '.png'
                                #img_filename = "{}/image_name{}-{}.png".format(location,current_page_index, img_index)   # если это чёрно-белое или цветное изображение  
                                if image.n < 5:
                                    image.writePNG(img_filename)                                                    #если это CMYK: конвертируем в RGB
                                else:                
                                    new_image = fitz.Pixmap(fitz.csRGB, image)
                                    new_image.writePNG(img_filename)
                                img = Image.open(img_filename)
                                file_name = img.filename
                                file_name = file_name.split(".")[0]
                                #print (str(file_name))
                                try:
                                    text = pytesseract.image_to_string(img, lang='rus').strip()
                                except Exception:
                                    print('ошибка распознования файла -' +str(file_name ))
                                row_links.append(text)
                                y = number_of_pages
                        while  y < 10:
                            row_links.append('      ') # добавляем в строку пустые элементы
                            y += 1
                else:  #   ячейка с ссылкой выбрана не верно
                    print('отсутствует ссылка')
                    row_links=['************','************','************','************','************','************','************','************','************']
                tab_link.append(row_links)
                nom_stroki = i-2
                if (nom_stroki//10)==(nom_stroki/10):  #десятая или сотая строка
                    print('десятая строка!!!')
                    df = pd.DataFrame(tab_link, columns=["номер БУ", "ссылка","листов","лист1","лист2","лист3","лист4","лист5","лист6","лист7","лист8","лист9","лист10"])
                    vr_file_name = 'c:/bu/tmp/resul'+str(nom_stroki)+'.xlsx'
                    element = tab_link.clear()
                    try:   
                        df.to_excel(vr_file_name)
                    except Exception:
                        print('внимание! ошибка записи в файл: '+ vr_file_name)
                
            
            df = pd.DataFrame(tab_link, columns=["номер БУ", "ссылка","листов","лист1","лист2","лист3","лист4","лист5","лист6","лист7","лист8","лист9","лист10"])
            try:   
                df.to_excel('c:/bu/tmp/resultat.xlsx')
            except Exception:
                print('внимание! ошибка записи в файл')
                print(tab_link)
            finally:
                print('успешное завершение записи в файл!')
        #   тут выхов функции объединяющей файлы
        #
    def onExit(self):
        global app
        concate_exel() #  вызываем объеденитель        
        dir_cls()  # очищаем директорию и выходим из кнопки 
        app.quit()
def choose_rb(event):
    global rb, ra
    rb = txt0.current()
    ra = txt1.current()
    #print (" выбраны колонки = "+ str(rb)+"  и "+ str(ra))
    

if __name__ == "__main__":
    app = App()
    app.title("Программа распознования сканов актов возврата БУ")
    app.mainloop()
