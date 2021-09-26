# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('c:/bu/resultat.xlsx')

# Get sheet names
print(wb.get_sheet_names())