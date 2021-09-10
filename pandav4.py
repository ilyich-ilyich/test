from openpyxl import load_workbook
import pandas as pd
import requests

df = pd.read_excel('c:/bu/1.xlsx')
wb = load_workbook('c:/bu/2.xlsx')
ws = wb.active
adressURL=ws.cell(row=6, column=9).hyperlink.target
headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/12.0.3 Safari/605.1.15",
            "Accept-Language": "en"
        }
send=requests.get(adressURL, headers=headers)
adressOS ='c:/bu/' +str(ws.cell(row=6, column=4).value)+'.pdf'
print(adressOS)
if send.status_code == 200:
    print('Success!')
    with open(adressOS,"wb")as code:
        code.write(send.content) #записываем результат

