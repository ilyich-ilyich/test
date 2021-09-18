from openpyxl import load_workbook
import pandas as pd
import requests

#df = pd.read_excel('c:/bu/1.xlsx')



wb = load_workbook('c:/bu/2.xlsx')
ws = wb.active
adressUrl =ws.cell(row=6, column=9).value
headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/12.0.3 Safari/605.1.15",
            "Accept-Language": "en"
            }
send=requests.get(adressUrl, headers=headers)
print(ws.cell(row=6, column=4).value)
adress ='c:/bu/' +str(ws.cell(row=6, column=4).value)+'.pdf'
print(adress)

with open(adress,"wb")as code:
    code.write(send.content)

links = []
for i in range(8, ws.max_row+1):  # 2nd arg in range() not inclusive, so add 1

    links.append([ws.cell(row=i, column=4).value, ws.cell(row=i, column=9).value])
print(links)    
df = pd.DataFrame(links, columns=["nomer_bu", "adress"])
df.to_excel('c:/bu/1.xlsx')
