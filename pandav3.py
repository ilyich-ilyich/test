from openpyxl import load_workbook
import pandas as pd
import requests

df = pd.read_excel('c:/bu/1.xlsx')



wb = load_workbook('c:/bu/2.xlsx')
ws = wb.active
#print(wb)
print(ws.cell(row=6, column=9).hyperlink.target)
send=requests.post(ws.cell(row=6, column=9).hyperlink.target)
file = open('D://bdseo.html','w')
print(ws.cell(row=6, column=4).value)
adress ='c:/bu/' +str(ws.cell(row=6, column=4).value)+'.pdf'
print(adress)
file = open(adress,'w')
file.write(send.text) #записываем результат
file.close()


links = []
for i in range(8, ws.max_row + 1):  # 2nd arg in range() not inclusive, so add 1
#    links.append(ws.cell(row=i, column=9).hyperlink.target)
    links.append(ws.cell(row=i, column=4).value)
    links.append(ws.cell(row=i, column=9).value)
    #print("i=", i)
#print(links)
#df['link'] = links