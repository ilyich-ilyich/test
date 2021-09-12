import tkinter as tk
import tkinter.filedialog as fd
import os
import sys
#import PyPDF2
import fitz
import pytesseract
from PIL import Image
from openpyxl import load_workbook
import pandas as pd
import requests


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        btn_file = tk.Button(self, text="Выбрать файл",
                             command=self.choose_file)
        btn_file.pack(padx=60, pady=10)
        btn_exit =tk.Button(self, text ="Выход", command=self.onExit)
        btn_exit.pack(padx=60, pady=10)

    def choose_file(self):
        filetypes = (("Exel файл", "*.xlsx"),
                     ("PDF файл", "*.pdf"),
                     ("Любой", "*"))
        filename = fd.askopenfilename(title="Открыть файл", initialdir="/",                                 # получили имя файла с путем
                                      filetypes=filetypes)
                                      
        if filename:
            print(filename)
            # работа с экселем
            tab_link = []                                                                                      # список строк для будущей таблицы   
            wb = load_workbook(filename)                                                                       # файл с ссылками на сканы актов возврата БУ
            ws = wb.active                                                                                     # выбор активной страницы
            headers = {
                        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/12.0.3 Safari/605.1.15",
                        "Accept-Language": "en"
                    }
            for i in range(4, ws.max_row + 1):                                                              # перебираем строки исх таблицы
                print("обрабатываем  "str(i)+ "  из "+ str(ws.max_row + 1))
                row_links =[] # строка для будущей таблицы    
                nom_bu = (ws.cell(row=i, column=4).value)                                                   # получаем номер БУ 
                adressURL=(ws.cell(row=i, column=9).value)                                                  # получаем ссылку
                send=requests.get(adressURL, headers=headers)                                               # загружаем файл в объект
                name_of_PDF_file ='c:/bu/tmp/' +str(nom_bu)+'.pdf'
                #print(name_of_PDF_file)
                if send.status_code == 200:                                                                 # если успешно  записывваем объект как файл pdf
                    #print('Success!')
                    row_links.append(nom_bu)                                                                #в список добавляем номер БУ
                    row_links.append(adressURL)                                                             #в список добавляемссылку на скан акта
                    with open(name_of_PDF_file,"wb")as code:
                        code.write(send.content)                                                                #записываем результат  в папку tmp
                        
                    # работа с pdf
                    pdf_file = fitz.open(name_of_PDF_file)                                                      #  открываем скачаный файл
                    location = 'c:/bu/tmp/'                                                                     #  папка для текстовых файлов
                    number_of_pages = len(pdf_file)
                    row_links.append(number_of_pages) #
                    #подсчет количества страниц в pdf
                    #print(row_links)
                    number_of_pages =1
                    for current_page_index in range(number_of_pages):                     #итерация по каждой странице в pdf
                        for img_index,img in enumerate(pdf_file.getPageImageList(current_page_index)):          # итерация по каждому изображению на каждой странице PDF
                            xref = img[0]
                            image = fitz.Pixmap(pdf_file, xref)
                            img_filename = name_of_PDF_file+'-'+str(current_page_index)+'-'+str(img_index) + '.png'
                            #img_filename = "{}/image_name{}-{}.png".format(location,current_page_index, img_index)   # если это чёрно-белое или цветное изображение  
                            if image.n < 5:
                                image.writePNG(img_filename)                                                    #если это CMYK: конвертируем в RGB
                            else:                
                                new_image = fitz.Pixmap(fitz.csRGB, image)
                                new_image.writePNG(img_filename)
                            img = Image.open(img_filename)
                            file_name = img.filename
                            file_name = file_name.split(".")[0]
                            #print (str(file_name))
                            text = pytesseract.image_to_string(img, lang='rus').strip()
                            row_links.append(text)
                            tab_link.append(row_links)
                            #with open(f'{file_name}.txt', 'w') as text_file:
                             #   text_file.write(text)
            df = pd.DataFrame(tab_link, columns=["nomer_bu", "adress","pages","text1","text2","text3"])
            df.to_excel('c:/bu/1.xlsx')
            
        print("все готово!")
    def onExit(self):
        global app
        app.quit()
    

if __name__ == "__main__":
    app = App()
    app.mainloop()
